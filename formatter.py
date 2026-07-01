from __future__ import annotations

import datetime as dt
import re
import struct
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


@dataclass
class Employee:
    employee_id: str
    name: str
    department: str = ""
    punches: list[str] | None = None


class XlsReader:
    END = 0xFFFFFFFE
    FREE = 0xFFFFFFFF

    def __init__(self, data: bytes):
        self.data = data
        self.sector_size = 512
        self.fat: list[int] = []
        self.workbook = b""
        self.sheets: list[tuple[str, int]] = []
        self.strings: list[str] = []
        self._read_ole()
        self._read_workbook()

    def _sector(self, index: int) -> bytes:
        start = (index + 1) * self.sector_size
        return self.data[start : start + self.sector_size]

    def _chain(self, start: int) -> list[int]:
        out: list[int] = []
        seen: set[int] = set()
        index = start
        while index not in (self.END, self.FREE) and index < 0xFFFFFFF0 and index not in seen:
            seen.add(index)
            out.append(index)
            index = self.fat[index]
        return out

    def _stream(self, start: int, size: int) -> bytes:
        return b"".join(self._sector(i) for i in self._chain(start))[:size]

    def _read_ole(self):
        header = self.data[:512]
        if header[:8] != b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
            raise ValueError("The uploaded file is not an old Excel .xls workbook.")

        self.sector_size = 1 << struct.unpack_from("<H", header, 30)[0]
        fat_count = struct.unpack_from("<I", header, 44)[0]
        dir_start = struct.unpack_from("<I", header, 48)[0]
        difat = list(struct.unpack_from("<109I", header, 76))
        fat_sectors = [i for i in difat[:fat_count] if i < 0xFFFFFFF0]
        for sector in fat_sectors:
            self.fat.extend(struct.unpack(f"<{self.sector_size // 4}I", self._sector(sector)))

        dir_bytes = self._stream(dir_start, self.sector_size * len(self._chain(dir_start)))
        workbook_entry = None
        for offset in range(0, len(dir_bytes), 128):
            entry = dir_bytes[offset : offset + 128]
            if len(entry) < 128:
                continue
            name_length = struct.unpack_from("<H", entry, 64)[0]
            if name_length < 2:
                continue
            name = entry[: name_length - 2].decode("utf-16le", "ignore")
            start = struct.unpack_from("<I", entry, 116)[0]
            size = struct.unpack_from("<Q", entry, 120)[0]
            if name in {"Workbook", "Book"}:
                workbook_entry = (start, size)
                break
        if not workbook_entry:
            raise ValueError("No Excel workbook stream was found in the upload.")
        self.workbook = self._stream(*workbook_entry)

    def _records(self, start=0):
        pos = start
        while pos + 4 <= len(self.workbook):
            record_id, length = struct.unpack_from("<HH", self.workbook, pos)
            payload = self.workbook[pos + 4 : pos + 4 + length]
            yield pos, record_id, payload
            pos += 4 + length

    def _read_workbook(self):
        sst_raw = b""
        for pos, record_id, payload in self._records():
            if record_id == 0x0085:
                sheet_offset = struct.unpack_from("<I", payload, 0)[0]
                name_len = payload[6]
                flags = payload[7]
                if flags & 1:
                    name = payload[8 : 8 + name_len * 2].decode("utf-16le", "ignore")
                else:
                    name = payload[8 : 8 + name_len].decode("latin1", "ignore")
                self.sheets.append((name, sheet_offset))
            elif record_id == 0x00FC:
                sst_raw += payload
                next_pos = pos + 4 + len(payload)
                while next_pos + 4 <= len(self.workbook):
                    next_id, next_len = struct.unpack_from("<HH", self.workbook, next_pos)
                    if next_id != 0x003C:
                        break
                    sst_raw += self.workbook[next_pos + 4 : next_pos + 4 + next_len]
                    next_pos += 4 + next_len
        self.strings = self._parse_sst(sst_raw)

    def _parse_sst(self, raw: bytes) -> list[str]:
        if len(raw) < 8:
            return []
        _, unique_count = struct.unpack_from("<II", raw, 0)
        strings: list[str] = []
        pos = 8
        for _ in range(unique_count):
            if pos + 3 > len(raw):
                break
            char_count = struct.unpack_from("<H", raw, pos)[0]
            pos += 2
            flags = raw[pos]
            pos += 1
            is_utf16 = bool(flags & 1)
            rich = bool(flags & 8)
            extended = bool(flags & 4)
            rich_runs = 0
            ext_len = 0
            if rich:
                rich_runs = struct.unpack_from("<H", raw, pos)[0]
                pos += 2
            if extended:
                ext_len = struct.unpack_from("<I", raw, pos)[0]
                pos += 4
            byte_count = char_count * (2 if is_utf16 else 1)
            text = raw[pos : pos + byte_count].decode("utf-16le" if is_utf16 else "latin1", "ignore")
            pos += byte_count + rich_runs * 4 + ext_len
            strings.append(text)
        return strings

    def sheet(self, name: str) -> dict[tuple[int, int], Any]:
        match = next((s for s in self.sheets if s[0] == name), None)
        if not match:
            raise ValueError(f"Sheet '{name}' was not found in the DTR file.")
        _, start = match
        cells: dict[tuple[int, int], Any] = {}
        for pos, record_id, payload in self._records(start):
            if record_id == 0x000A and pos > start:
                break
            try:
                if record_id == 0x00FD and len(payload) >= 10:
                    row, col, _, idx = struct.unpack_from("<HHHI", payload, 0)
                    cells[(row + 1, col + 1)] = self.strings[idx] if idx < len(self.strings) else ""
                elif record_id == 0x0204 and len(payload) >= 8:
                    row, col, _, length = struct.unpack_from("<HHHH", payload, 0)
                    cells[(row + 1, col + 1)] = payload[8 : 8 + length].decode("latin1", "ignore")
                elif record_id == 0x00D6 and len(payload) >= 8:
                    row, col, _, length = struct.unpack_from("<HHHH", payload, 0)
                    cells[(row + 1, col + 1)] = payload[8 : 8 + length * 2].decode("utf-16le", "ignore")
                elif record_id == 0x0203 and len(payload) >= 14:
                    row, col, _ = struct.unpack_from("<HHH", payload, 0)
                    value = struct.unpack_from("<d", payload, 6)[0]
                    cells[(row + 1, col + 1)] = value
            except (IndexError, struct.error):
                continue
        return cells


def convert_dtr(input_path: Path, original_name: str, output_dir: Path) -> tuple[Path, dict[str, Any]]:
    data = input_path.read_bytes()
    reader = XlsReader(data)
    employees = extract_employees(reader)
    if not employees:
        raise ValueError("No employee records were found in the DTR file.")

    dates = extract_dates(reader)
    if not dates:
        raise ValueError("No DTR cut-off dates were found in the file.")

    workbook = build_output(employees, dates)
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_stem = re.sub(r"[^A-Za-z0-9_-]+", "_", Path(original_name).stem).strip("_") or "dtr"
    output_path = output_dir / f"{safe_stem}_relayouted_{uuid.uuid4().hex[:8]}.xlsx"
    workbook.save(output_path)
    return output_path, {"employees": len(employees), "days": len(dates)}


def extract_employees(reader: XlsReader) -> list[Employee]:
    stats = reader.sheet("Att. Stat.")
    logs = reader.sheet("Att.log report")
    employees: list[Employee] = []
    row = 5
    while (row, 1) in stats or (row, 2) in stats:
        employee_id = clean(stats.get((row, 1)))
        name = clean(stats.get((row, 2)))
        department = clean(stats.get((row, 3)))
        if employee_id and name and employee_id != "1":
            employees.append(Employee(employee_id, name, department, []))
        row += 1

    by_id = {employee.employee_id: employee for employee in employees}
    max_row = max((r for r, _ in logs), default=0)
    for row in range(1, max_row + 1):
        if clean(logs.get((row, 1))) == "ID:":
            employee_id = clean(logs.get((row, 3)))
            employee = by_id.get(employee_id)
            if not employee:
                continue
            punch_row = row + 1
            punch_cols = sorted({col for (r, col) in logs if r == punch_row and col > 0})
            if punch_cols:
                max_col = max(punch_cols)
                employee.punches = [clean(logs.get((punch_row, col))) for col in range(1, max_col + 1)]
            else:
                employee.punches = []
    return [employee for employee in employees if any(employee.punches or [])]


def extract_dates(reader: XlsReader) -> list[dt.date]:
    for sheet_name in ("Exception Stat.", "Att.log report", "Schedule Infor."):
        cells = reader.sheet(sheet_name)
        values = [clean(value) for value in cells.values()]
        for value in values:
            match = re.search(r"(20\d\d-\d\d-\d\d)\s*~\s*(20\d\d-\d\d-\d\d)", value)
            if match:
                start = dt.date.fromisoformat(match.group(1))
                end = dt.date.fromisoformat(match.group(2))
                return [start + dt.timedelta(days=i) for i in range((end - start).days + 1)]
    return []


def build_output(employees: list[Employee], dates: list[dt.date]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.sheet_view.showGridLines = True

    widths = {
        "A": 12.86, "B": 11.71, "C": 17.71, "D": 9.14, "E": 9.14,
        "F": 9.14, "G": 6.57, "H": 11.43, "I": 7.57, "J": 11.43,
        "K": 9.14, "L": 12.0, "M": 8.0, "N": 9.14, "O": 14.71,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    start_row = 1
    for employee in employees:
        write_block(ws, start_row, employee, dates)
        start_row += 34

    ws.print_area = f"A1:O{start_row - 3}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return wb


def write_block(ws, row: int, employee: Employee, dates: list[dt.date]):
    title_font = Font(name="Calibri", size=11, bold=False)
    header_font = Font(name="Calibri", size=11, bold=True)
    font = Font(name="Calibri", size=11)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    alert_fill = PatternFill("solid", fgColor="FF0000")

    ws.row_dimensions[row].height = 20.25
    ws.cell(row, 1, "Attendance Record Report").font = title_font
    ws.cell(row + 1, 1, "Name ")
    ws.cell(row + 1, 2, format_name(employee.name))
    ws.cell(row + 2, 1, "ID No.")
    ws.cell(row + 2, 2, employee.employee_id)
    ws.cell(row + 2, 10, "Cut Off Period")
    ws.cell(row + 2, 11, cutoff_label(dates))

    headers = [
        "Date", "Days", "Schedule", "IN", "OUT", "# of days", "OT (HRS)",
        "RD OT (HRS)", "RH OT", "Late (Mins.)", "Night Dif.", "OT Night Diff",
        "Leave", "Remarks", "",
    ]
    header_row = row + 4
    ws.row_dimensions[header_row].height = 36
    for col, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col, header)
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for index, day in enumerate(dates):
        excel_row = header_row + 1 + index
        punch = (employee.punches or [""] * len(dates))[index] if index < len(employee.punches or []) else ""
        in_time, out_time = split_punch(punch)
        is_holiday = day.day == 13
        is_rest = day.weekday() == 6
        schedule = schedule_for(day, in_time, is_holiday, employee.department)
        display_in = in_time or "00:00"
        display_out = out_time or "00:00"
        worked = 1 if in_time and out_time and not is_rest and not is_holiday else ""
        ot = overtime_hours(out_time, schedule) if worked else ""
        rd_ot = rest_day_hours(in_time, out_time) if in_time and out_time and is_rest else ""
        rh_ot = rest_day_hours(in_time, out_time) if in_time and out_time and is_holiday else ""
        late = late_minutes(in_time, schedule) if in_time and not is_rest and not is_holiday else ""
        leave = 1 if not in_time and not out_time and not is_rest and not is_holiday else ""
        remarks = "HOLIDAY" if is_holiday else ""

        values = [
            day, day.strftime("%A"), schedule, display_in, display_out, worked, ot,
            rd_ot, rh_ot, late, "", "", leave, remarks, "",
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(excel_row, col, value)
            cell.font = font
            cell.alignment = center if col not in (3, 14, 15) else left
            if col == 1:
                cell.number_format = "mmmm d, yyyy"
            if col == 4 and late:
                cell.fill = alert_fill
            if col == 5 and in_time and not out_time:
                cell.fill = alert_fill
            cell.border = border
        ws.merge_cells(start_row=excel_row, start_column=14, end_row=excel_row, end_column=15)

    first_date_row = header_row + 1
    last_date_row = first_date_row + len(dates) - 1
    total_row = last_date_row + 1
    for col in range(1, 16):
        ws.cell(total_row, col).border = border
        ws.cell(total_row, col).font = font
        ws.cell(total_row, col).alignment = center if col not in (14, 15) else left
    for col in range(6, 14):
        letter = get_column_letter(col)
        ws.cell(total_row, col, f"=SUM({letter}{first_date_row}:{letter}{last_date_row})")
    ws.cell(total_row, 14, f'=COUNTIF(N{first_date_row}:O{last_date_row},"HOLIDAY")')
    ws.merge_cells(start_row=total_row, start_column=14, end_row=total_row, end_column=15)

    labels = [
        ("Reg. Work Days", f"=+F{total_row}"),
        ("Leave:", f"=M{total_row}"),
        ("Late (mins.):", f"=J{total_row}"),
        ("REG Holiday:" if row == 1 else "Holiday:", f'=COUNTIF(N{first_date_row}:O{last_date_row},"HOLIDAY")'),
        ("Regular OT:", f"=G{total_row}"),
        ("RD / Sunday OT:", f"=H{total_row}"),
        ("Reg Hol OT:", f"=I{total_row}"),
        ("Night Differential:", f"=K{total_row}"),
        ("OT Night Differential:", f"=L{total_row}"),
        ("Sunday Night Differential:", ""),
        ("Sunday OT Night Differential:", ""),
    ]
    for i, (label, value) in enumerate(labels, 1):
        r = total_row + i
        ws.cell(r, 1, label).font = font
        ws.cell(r, 4, value).font = font
        ws.cell(r, 4).alignment = center


def split_punch(value: str) -> tuple[str, str]:
    parts = re.findall(r"\d{1,2}:\d{2}", value or "")
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], parts[-1]


def schedule_for(day: dt.date, in_time: str, is_holiday: bool, department: str) -> str:
    if is_holiday:
        return "NO WORK"
    if day.weekday() == 6:
        return "RD"
    if department in {"Management", "A&F", "QC", "WHSE", "Utility"}:
        return "8:00-5:00"
    hour = int(in_time.split(":", 1)[0]) if in_time else 8
    return "6:00am-3:00pm" if hour < 7 else "8:00-5:00"


def overtime_hours(out_time: str, schedule: str) -> int | str:
    if not out_time:
        return ""
    hour, minute = [int(part) for part in out_time.split(":")]
    end_hour = 15 if schedule.startswith("6:00") else 17
    diff = (hour * 60 + minute) - end_hour * 60
    return max(0, round(diff / 60)) or ""


def rest_day_hours(in_time: str, out_time: str) -> int | str:
    start = minutes_since_midnight(in_time)
    end = minutes_since_midnight(out_time)
    if start is None or end is None:
        return ""
    if end < start:
        end += 24 * 60
    return max(0, round((end - start) / 60)) or ""


def late_minutes(in_time: str, schedule: str) -> int | str:
    start = schedule_start_minutes(schedule)
    actual = minutes_since_midnight(in_time)
    if start is None or actual is None:
        return ""
    return max(0, actual - start) or ""


def schedule_start_minutes(schedule: str) -> int | None:
    if schedule.startswith("6:00"):
        return 6 * 60
    if schedule.startswith("8:00"):
        return 8 * 60
    return None


def minutes_since_midnight(value: str) -> int | None:
    if not value:
        return None
    match = re.fullmatch(r"(\d{1,2}):(\d{2})", value)
    if not match:
        return None
    hour, minute = int(match.group(1)), int(match.group(2))
    return hour * 60 + minute


def format_name(name: str) -> str:
    parts = name.strip().split()
    if len(parts) < 2:
        return name
    return f"{parts[-1]}, {' '.join(parts[:-1])}"


def cutoff_label(dates: list[dt.date]) -> str:
    if not dates:
        return ""
    start, end = dates[0], dates[-1]
    month = start.strftime("%b")
    if start.month == end.month and start.year == end.year:
        return f"{month} {start.day}-{end.day}, {end.year}"
    return f"{start:%b %-d, %Y} - {end:%b %-d, %Y}"


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\x0c", "").strip()
    if re.fullmatch(r"[+-]?\d+\.\d+", text):
        try:
            number = float(text)
            if abs(number) < 1e-250:
                return ""
        except ValueError:
            pass
    return text
